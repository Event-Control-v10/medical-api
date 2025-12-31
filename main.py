import os
import io
import time
import base64
import json
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import load_workbook
from groq import Groq
from PIL import Image

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

HISTORY_DIR = "history"
if not os.path.exists(HISTORY_DIR): os.makedirs(HISTORY_DIR)

client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# Modèles
MODEL_EXCEL = "llama-3.3-70b-versatile"
MODEL_VISION = "llama-3.2-11b-vision-preview"

@app.get("/healthz")
async def healthz(): return {"status": "ok"}

# --- PARTIE EXCEL (Inchangée car elle marche) ---
@app.post("/process")
async def process_excel(file: UploadFile = File(...), instruction: str = Form(None), audio: UploadFile = File(None)):
    try:
        file_bytes = await file.read()
        df_orig = pd.read_excel(io.BytesIO(file_bytes), skiprows=3)
        df_mod = df_orig.copy()

        user_text = instruction
        if audio:
            audio_data = await audio.read()
            trans = client.audio.transcriptions.create(file=("a.wav", audio_data), model="whisper-large-v3", language="fr")
            user_text = trans.text

        if not user_text: return {"error": "Aucune consigne"}

        prompt = f"""
        DataFrame df: {list(df_orig.columns)}. Instruction: {user_text}. 
        Code Python uniquement (df.at[index, 'col'] = val). Pas de blabla.
        """
        
        chat = client.chat.completions.create(model=MODEL_EXCEL, messages=[{"role": "user", "content": prompt}], temperature=0)
        code = chat.choices[0].message.content.strip().replace("```python", "").replace("```", "")

        exec_scope = {"df": df_mod, "pd": pd}
        exec(code, {}, exec_scope)
        df_mod = exec_scope["df"]

        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        START_ROW = 5 
        
        for r_idx, row in enumerate(df_mod.values):
            for c_idx, val in enumerate(row):
                ws.cell(row=START_ROW + r_idx, column=c_idx + 1).value = val

        ts = time.strftime("%Y%m%d-%H%M%S")
        fname = f"Reginalde_{ts}.xlsx"
        fpath = os.path.join(HISTORY_DIR, fname)
        wb.save(fpath)
        return FileResponse(fpath, filename=fname)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- PARTIE SCANNER GROQ (Correction Anti-Argumentation) ---
@app.post("/ocr")
async def ocr(image: UploadFile = File(...)):
    try:
        img_bytes = await image.read()
        
        # 1. Traitement image (Sécurité)
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode != "RGB":
            img = img.convert("RGB")
        img.thumbnail((1024, 1024))
        
        buffered = io.BytesIO()
        img.save(buffered, format="JPEG")
        base64_image = base64.b64encode(buffered.getvalue()).decode('utf-8')

        # 2. PROMPT STRICT "JSON ONLY"
        # On force l'IA à répondre en JSON pur, ce qui l'empêche de "parler".
        messages = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": """
                        Analyse cette image.
                        Extrait le contenu textuel complet.
                        
                        Ta réponse doit être UNIQUEMENT un objet JSON brut contenant une seule clé "content".
                        Ne dis pas "Voici le texte" ou "L'image contient".
                        Juste le JSON.
                        Exemple format: {"content": "Titre du doc\nNom du patient..."}
                        """
                    },
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                ]
            }
        ]

        # 3. Appel Groq avec réponse forcée en JSON
        response = client.chat.completions.create(
            model=MODEL_VISION,
            messages=messages,
            temperature=0, # Zéro créativité = Zéro blabla
            response_format={"type": "json_object"} # FORCE LE SILENCE
        )
        
        # 4. On récupère le contenu propre
        json_content = json.loads(response.choices[0].message.content)
        clean_text = json_content.get("content", "Aucun texte détecté.")

        return {
            "text": clean_text,
            "image": f"data:image/jpeg;base64,{base64_image}"
        }

    except Exception as e:
        print(f"ERREUR OCR : {str(e)}")
        # Fallback au cas où le modèle JSON échoue, on renvoie l'erreur proprement
        raise HTTPException(status_code=500, detail=f"Erreur Lecture: {str(e)}")

@app.get("/history")
async def get_history():
    files = sorted(os.listdir(HISTORY_DIR), reverse=True)[:15]
    return {"files": files}

@app.get("/download/{filename}")
async def download(filename: str):
    return FileResponse(os.path.join(HISTORY_DIR, filename))
